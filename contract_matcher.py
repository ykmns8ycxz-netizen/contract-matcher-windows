import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
from pathlib import Path
import shutil
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
import warnings
import tempfile
import subprocess
import platform
warnings.filterwarnings('ignore')

class ContractMatcherApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("南银理财 - 合同文件匹配工具")
        self.root.geometry("1200x900")  # 进一步增大窗口尺寸
        self.root.configure(bg='#f0f0f0')
        
        # 存储文件路径
        self.excel_file = None
        self.pdf_files = []
        self.output_dir = None
        
        self.setup_ui()
        
    def setup_ui(self):
        """设置用户界面"""
        # 主容器 - 使用grid布局确保所有元素可见
        main_container = tk.Frame(self.root, bg='#ffffff', relief='raised', bd=2)
        main_container.pack(fill='both', expand=True, padx=30, pady=30)
        
        # 配置主容器的grid权重
        main_container.grid_rowconfigure(1, weight=1)  # 状态区域可扩展
        main_container.grid_rowconfigure(2, weight=0)  # 按钮区域固定
        main_container.grid_rowconfigure(3, weight=0)  # 版本信息固定
        main_container.grid_columnconfigure(0, weight=1)
        
        # 标题栏
        header_frame = tk.Frame(main_container, bg='#1e3d6f', height=120)
        header_frame.grid(row=0, column=0, sticky='ew', pady=(0, 20))
        header_frame.grid_propagate(False)
        
        # 银行logo和标题 - 进一步增大字体
        logo_label = tk.Label(header_frame, text="南银理财", 
                             font=("Arial", 24, "bold"),
                             fg='white', bg='#1e3d6f',
                             justify='center')
        logo_label.pack(expand=True, pady=15)
        
        subtitle_label = tk.Label(header_frame, text="合同文件智能匹配系统", 
                                 font=("Arial", 16),
                                 fg='#cccccc', bg='#1e3d6f')
        subtitle_label.pack(pady=(0, 15))
        
        # 文件上传区域 - 使用网格布局
        upload_frame = tk.Frame(main_container, bg='#ffffff')
        upload_frame.grid(row=1, column=0, sticky='ew', pady=(0, 20))
        
        # Excel文件上传区域
        excel_frame = tk.LabelFrame(upload_frame, text=" 📊 Excel合同台账文件 ", 
                                   font=("Arial", 14, "bold"),
                                   bg='#ffffff', fg='#1e3d6f',
                                   relief='groove', bd=2, padx=20, pady=20)
        excel_frame.grid(row=0, column=0, sticky='ew', padx=(0, 15))
        
        self.excel_label = tk.Label(excel_frame, text="请选择Excel文件...", 
                                   font=("Arial", 12),
                                   fg='#666666', bg='#ffffff', 
                                   wraplength=400, justify='left')
        self.excel_label.pack(side='left', fill='x', expand=True)
        
        excel_btn = tk.Button(excel_frame, text="选择文件", 
                             font=("Arial", 12),
                             command=self.select_excel_file,
                             bg='#1e3d6f', fg='white',
                             relief='raised', bd=2, padx=20, pady=8)
        excel_btn.pack(side='right', padx=(15, 0))
        
        # PDF文件上传区域
        pdf_frame = tk.LabelFrame(upload_frame, text=" 📑 PDF合同文件 ", 
                                 font=("Arial", 14, "bold"),
                                 bg='#ffffff', fg='#1e3d6f',
                                 relief='groove', bd=2, padx=20, pady=20)
        pdf_frame.grid(row=0, column=1, sticky='ew', padx=(15, 0))
        
        self.pdf_label = tk.Label(pdf_frame, text="请选择PDF文件...", 
                                 font=("Arial", 12),
                                 fg='#666666', bg='#ffffff',
                                 wraplength=400, justify='left')
        self.pdf_label.pack(side='left', fill='x', expand=True)
        
        pdf_btn = tk.Button(pdf_frame, text="选择文件", 
                           font=("Arial", 12),
                           command=self.select_pdf_files,
                           bg='#1e3d6f', fg='white',
                           relief='raised', bd=2, padx=20, pady=8)
        pdf_btn.pack(side='right', padx=(15, 0))
        
        # 配置网格权重
        upload_frame.columnconfigure(0, weight=1)
        upload_frame.columnconfigure(1, weight=1)
        
        # 状态显示区域
        status_frame = tk.LabelFrame(main_container, text=" 🔍 处理状态 ", 
                                    font=("Arial", 14, "bold"),
                                    bg='#ffffff', fg='#1e3d6f',
                                    relief='groove', bd=2, padx=20, pady=20)
        status_frame.grid(row=2, column=0, sticky='nsew', pady=(0, 20))
        
        # 创建状态文本框架
        text_frame = tk.Frame(status_frame, bg='#f8f9fa')
        text_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.status_text = tk.Text(text_frame, wrap="word",
                                  font=("Arial", 12),
                                  bg='#f8f9fa', fg='#333333',
                                  relief='flat', padx=15, pady=15,
                                  height=8)  # 固定高度
        self.status_text.pack(side='left', fill='both', expand=True)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(text_frame, command=self.status_text.yview)
        scrollbar.pack(side='right', fill='y')
        self.status_text.config(yscrollcommand=scrollbar.set)
        
        # 按钮区域 - 确保在状态区域下方
        button_frame = tk.Frame(main_container, bg='#ffffff')
        button_frame.grid(row=3, column=0, sticky='ew', pady=(0, 15))
        
        # 确认按钮
        confirm_btn = tk.Button(button_frame, text="🚀 开始匹配处理", 
                               font=("Arial", 15, "bold"),
                               command=self.process_files,
                               bg='#d32f2f', fg='white',
                               relief='raised', bd=3, padx=40, pady=12)
        confirm_btn.pack(pady=10)
        
        # 重置按钮
        reset_btn = tk.Button(button_frame, text="🔄 重置重新上传", 
                             font=("Arial", 13, "bold"),
                             command=self.reset_files,
                             bg='#1e3d6f', fg='white',
                             relief='raised', bd=2, padx=30, pady=10)
        reset_btn.pack(pady=5)
        
        # 版本信息
        version_label = tk.Label(main_container, text="南银理财 © 2025 - 合同匹配工具 v1.0", 
                                font=("Arial", 11),
                                fg='#999999', bg='#ffffff')
        version_label.grid(row=4, column=0, sticky='ew', pady=10)
        
    def select_excel_file(self):
        """选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel合同台账文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.excel_file = file_path
            display_name = os.path.basename(file_path)
            # 如果文件名太长，截断显示
            if len(display_name) > 30:
                display_name = display_name[:27] + "..."
            self.excel_label.config(text=display_name, fg='#1e3d6f')
            self.log_status(f"✅ 已选择Excel文件: {os.path.basename(file_path)}")
            
    def select_pdf_files(self):
        """选择PDF文件"""
        file_paths = filedialog.askopenfilenames(
            title="选择PDF合同文件",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_paths:
            self.pdf_files = list(file_paths)
            file_names = [os.path.basename(f) for f in file_paths]
            display_text = f"已选择 {len(file_paths)} 个PDF文件"
            self.pdf_label.config(text=display_text, fg='#1e3d6f')
            self.log_status(f"✅ {display_text}")
            # 显示所有文件名
            for i, name in enumerate(file_names):
                if len(name) > 40:  # 如果文件名太长，截断显示
                    display_name = name[:37] + "..."
                else:
                    display_name = name
                self.log_status(f"   📄 {display_name}")
            
    def reset_files(self):
        """重置所有文件选择"""
        self.excel_file = None
        self.pdf_files = []
        self.output_dir = None
        
        # 重置界面显示
        self.excel_label.config(text="请选择Excel文件...", fg='#666666')
        self.pdf_label.config(text="请选择PDF文件...", fg='#666666')
        
        # 清空状态信息
        self.clear_status()
        self.log_status("🔄 文件选择已重置，可以重新上传文件")
        self.log_status("请重新选择Excel文件和PDF文件")
        
    def log_status(self, message):
        """记录状态信息"""
        self.status_text.insert("end", message + "\n")
        self.status_text.see("end")
        self.root.update()
        
    def clear_status(self):
        """清空状态信息"""
        self.status_text.delete(1.0, "end")
        
    def parse_pdf_filename(self, filename):
        """
        解析PDF文件名，提取机构、合同类型、合同编号
        格式: 机构-合同类型-合同编号.pdf
        """
        # 移除文件扩展名
        name_without_ext = os.path.splitext(filename)[0]
        
        # 分割文件名
        parts = name_without_ext.split('-')
        
        if len(parts) < 3:
            return None, None, None
            
        # 机构是第一个部分
        institution = parts[0]
        
        # 合同类型是中间部分（可能有多个-连接）
        contract_type = '-'.join(parts[1:-1])
        
        # 合同编号是最后一个部分
        contract_number = parts[-1]
        
        return institution, contract_type, contract_number
        
    def open_pdf_file(self, pdf_path):
        """打开PDF文件"""
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', pdf_path))
            elif platform.system() == 'Windows':  # Windows
                os.startfile(pdf_path)
            else:  # Linux
                subprocess.call(('xdg-open', pdf_path))
            return True
        except Exception as e:
            self.log_status(f"   ❌ 打开PDF文件失败: {str(e)}")
            return False

    def process_files(self):
        """处理文件匹配"""
        if not self.excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
            
        if not self.pdf_files:
            messagebox.showerror("错误", "请先选择PDF文件")
            return
            
        try:
            self.clear_status()
            self.log_status("🚀 开始处理文件匹配...")
            self.log_status("=" * 60)
            
            # 读取Excel文件获取列索引
            self.log_status("📊 读取Excel文件结构...")
            df_original = pd.read_excel(self.excel_file)
            
            # 检查必要的列是否存在
            required_columns = ['机构', '合同类型', '合同编号', '合同原件']
            missing_columns = [col for col in required_columns if col not in df_original.columns]
            if missing_columns:
                messagebox.showerror("错误", f"Excel文件中缺少必要的列: {', '.join(missing_columns)}")
                return
            
            # 获取列索引
            col_indices = {col: df_original.columns.get_loc(col) + 1 for col in required_columns}
            
            self.log_status(f"📈 Excel文件包含 {len(df_original)} 行数据")
            
            # 创建临时工作目录
            temp_dir = tempfile.mkdtemp(prefix="contract_temp_")
            
            # 复制原始Excel文件到临时位置进行处理
            temp_excel_path = os.path.join(temp_dir, "temp_workbook.xlsx")
            shutil.copy2(self.excel_file, temp_excel_path)
            
            # 解析PDF文件
            pdf_mapping = {}
            self.log_status("\n📑 解析PDF文件...")
            for pdf_path in self.pdf_files:
                filename = os.path.basename(pdf_path)
                temp_pdf_path = os.path.join(temp_dir, filename)
                shutil.copy2(pdf_path, temp_pdf_path)
                
                # 解析文件名
                institution, contract_type, contract_number = self.parse_pdf_filename(filename)
                if institution and contract_type and contract_number:
                    key = (institution, contract_type)
                    pdf_mapping[key] = {
                        'contract_number': contract_number,
                        'pdf_path': temp_pdf_path,
                        'filename': filename,
                        'relative_path': f"合同PDF附件/{filename}"  # 相对路径用于超链接
                    }
                    self.log_status(f"   ✅ 解析: {institution} - {contract_type} - {contract_number}")
                else:
                    self.log_status(f"   ❌ 无法解析: {filename}")
            
            self.log_status(f"\n✅ 成功解析 {len(pdf_mapping)} 个PDF文件")
            
            # 匹配并更新Excel
            matched_count = 0
            self.log_status("\n🔗 开始匹配数据并更新Excel...")
            
            # 使用openpyxl直接修改Excel，保持原有样式
            workbook = openpyxl.load_workbook(temp_excel_path)
            sheet = workbook.active
            
            for row_num, row in enumerate(df_original.iterrows(), 2):  # 从第2行开始（跳过标题行）
                institution = str(row[1]['机构']).strip()
                contract_type = str(row[1]['合同类型']).strip()
                
                key = (institution, contract_type)
                if key in pdf_mapping:
                    pdf_info = pdf_mapping[key]
                    
                    # 更新合同编号列
                    contract_num_cell = sheet.cell(row=row_num, column=col_indices['合同编号'])
                    contract_num_cell.value = pdf_info['contract_number']
                    
                    # 在合同原件列创建超链接
                    attachment_cell = sheet.cell(row=row_num, column=col_indices['合同原件'])
                    display_text = f"📎 {pdf_info['filename']}"
                    attachment_cell.value = display_text
                    
                    # 设置超链接（使用相对路径）
                    attachment_cell.hyperlink = pdf_info['relative_path']
                    attachment_cell.style = "Hyperlink"
                    
                    matched_count += 1
                    self.log_status(f"   ✅ 匹配成功: {institution} - {contract_type}")
                else:
                    # 显示未匹配的项目
                    if institution and contract_type and institution != 'nan' and contract_type != 'nan':
                        self.log_status(f"   ⚠️  未匹配: {institution} - {contract_type}")
            
            # 保存更新后的Excel
            workbook.save(temp_excel_path)
            
            # 让用户选择保存位置
            self.log_status("\n💾 准备保存结果...")
            save_path = filedialog.asksaveasfilename(
                title="保存更新后的Excel文件",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="合同台账_已匹配.xlsx"
            )
            
            if not save_path:  # 用户取消了保存
                self.log_status("❌ 用户取消保存操作")
                # 清理临时文件
                shutil.rmtree(temp_dir)
                return
            
            # 复制最终的Excel文件到用户选择的位置
            shutil.copy2(temp_excel_path, save_path)
            
            # 创建PDF附件文件夹（与Excel文件同级）
            save_dir = os.path.dirname(save_path)
            pdf_output_dir = os.path.join(save_dir, "合同PDF附件")
            os.makedirs(pdf_output_dir, exist_ok=True)
            
            # 复制所有PDF文件到附件文件夹
            copied_pdf_count = 0
            for pdf_info in pdf_mapping.values():
                try:
                    shutil.copy2(pdf_info['pdf_path'], os.path.join(pdf_output_dir, pdf_info['filename']))
                    copied_pdf_count += 1
                except Exception as e:
                    self.log_status(f"   ❌ 复制PDF失败: {pdf_info['filename']} - {str(e)}")
            
            # 记录输出目录，用于后续打开文件
            self.output_dir = save_dir
            
            # 生成处理报告
            report = f"""
🎉 处理完成！
{'=' * 60}
📊 数据统计:
   • 总数据行数: {len(df_original)}
   • 成功匹配: {matched_count}
   • 未匹配: {len(df_original) - matched_count}
   • PDF附件文件: {copied_pdf_count} 个

💾 输出文件:
   • 更新后的Excel(保持原样式): {os.path.basename(save_path)}
   • PDF附件目录: 合同PDF附件/

📍 保存位置: {save_dir}

📝 使用说明:
   • Excel文件保持了原有的所有格式和样式
   • 合同编号已自动更新
   • 合同原件列已创建可点击的超链接
   • 点击PDF文件名即可打开对应的合同文件
   • 所有PDF文件保存在"合同PDF附件"文件夹中
   • 点击"重置重新上传"按钮可以重新开始新的匹配任务
            """
            
            self.log_status(report)
            
            # 清理临时文件
            shutil.rmtree(temp_dir)
            
            # 显示完成消息并提供打开选项
            result = messagebox.askyesno("处理完成", 
                               f"🎉 文件处理完成！\n\n"
                               f"📊 匹配结果:\n"
                               f"   • 成功匹配: {matched_count} 个合同\n"
                               f"   • 未匹配: {len(df_original) - matched_count} 个\n"
                               f"   • PDF附件: {copied_pdf_count} 个\n\n"
                               f"💾 保存位置:\n{save_dir}\n\n"
                               f"是否打开输出文件夹？")
            
            if result:
                self.open_output_folder()
            
        except Exception as e:
            error_msg = f"❌ 处理过程中出现错误: {str(e)}"
            self.log_status(error_msg)
            messagebox.showerror("错误", error_msg)
    
    def open_output_folder(self):
        """打开输出文件夹"""
        if self.output_dir and os.path.exists(self.output_dir):
            try:
                if platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', self.output_dir))
                elif platform.system() == 'Windows':  # Windows
                    os.startfile(self.output_dir)
                else:  # Linux
                    subprocess.call(('xdg-open', self.output_dir))
            except Exception as e:
                self.log_status(f"❌ 打开文件夹失败: {str(e)}")
            
    def run(self):
        """运行应用程序"""
        self.root.mainloop()

if __name__ == "__main__":
    app = ContractMatcherApp()
    app.run()